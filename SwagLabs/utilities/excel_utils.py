import openpyxl
import xlrd

def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, col_num).value

def write_data(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, col_num).value = data
    workbook.save(file)

# def read_locators(sheet_name):
#     workbook = xlrd.open_workbook(r'C:\Users\User\PycharmProjects\SwagLabs\test_data\user_login.xlsx')
#     worksheet = workbook.sheet_by_name(sheet_name)
#
#     rows = worksheet.get_rows()
#     headers = next(rows) # skipping header
#
#     return {row[0].value: (row[1], row[2]) for row in rows}
#
# print(read_locators('user_login'))


